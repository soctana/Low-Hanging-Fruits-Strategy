import pandas as pd
import yfinance as yf
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# 抓取收盤價的函數
def get_additional_info(stock_code, market_type):
    if market_type == "上市":
        stock = yf.Ticker(f"{stock_code}.TW")
    elif market_type == "上櫃":
        stock = yf.Ticker(f"{stock_code}.TWO")

    last_close = '無資料'
    periods = ["1d", "5d", "1mo"]
    for period in periods:
        history_data = stock.history(period=period)
        if not history_data.empty:
            last_close = round(history_data['Close'].iloc[-1], 2)
            break

    return last_close

# 合併資料並抓取最新收盤價
def fetch_closing_prices():
    # 定義所有要讀取的檔案路徑
    file_paths = [
        r'G:\我的雲端硬碟\Horizon\python_stock\qualified_stocks_financial_data_A.xlsx',
        r'G:\我的雲端硬碟\Horizon\python_stock\qualified_stocks_financial_data_B.xlsx',
        r'G:\我的雲端硬碟\Horizon\python_stock\qualified_stocks_financial_data_C.xlsx',
        r'G:\我的雲端硬碟\Horizon\python_stock\qualified_stocks_financial_data_D.xlsx'
    ]

    # 讀取所有 Excel 檔案並合併成一個 DataFrame
    df_list = [pd.read_excel(file) for file in file_paths]
    df = pd.concat(df_list, ignore_index=True)

    # 針對每個股票抓取最新收盤價
    df['最新收盤價'] = df.apply(lambda row: get_additional_info(row['股票代碼'], row['市場類型']), axis=1)

    return df

# 計算與合併資料
def calculate_and_combine(df):
    # 讀取 "手動List" 資料
    manual_list_path = r'G:\我的雲端硬碟\Horizon\Additional Data_LHF.xlsx'
    manual_list_df = pd.read_excel(manual_list_path, sheet_name='手動List')

    # 比對 "股票代碼" 和 "公司代號"，並將 "手動List" 中的相應列加入 df
    df = df.merge(
        manual_list_df[['公司代號', 'Next EPS', 'EPS', '配息率', '下次配息時間', '下次配息金額', 'support', 'memo']],
        how='left',
        left_on='股票代碼',
        right_on='公司代號'
    )

    # 刪除 "公司代號" 欄位
    df.drop(columns=['公司代號'], inplace=True)

    df['前一次除息日'] = pd.to_datetime(df['前一次除息日'], format='%Y-%m-%d', errors='coerce').dt.date
    df['下一次除息日'] = pd.to_datetime(df['下一次除息日'], format='%Y-%m-%d', errors='coerce').dt.date

    # 判斷「前一次除息日」和「下一次除息日」的值是否相同，如果相同，清空「下一次除息日」和「下一次除息金額」欄位
    df.loc[df['前一次除息日'] == df['下一次除息日'], ['下一次除息日', '下一次除息金額']] = [pd.NaT, np.nan]

    # 新增 NDD 欄位並根據規則填入值
    df['NDD'] = np.where(df['下次配息時間'].notna(), df['下次配息時間'], 
                np.where(df['下一次除息日'].notna(), df['下一次除息日'], 
                np.where(df['前一次除息日'].notna(), df['前一次除息日'] + pd.DateOffset(days=365), pd.NaT)))

    # 將NDD欄位的數據轉換為日期格式並去除時間
    df['NDD'] = pd.to_datetime(df['NDD']).dt.date

    # 如果NDD的日期小於今天，則NDD加365天
    today = datetime.today().date()
    df['NDD'] = df['NDD'].apply(lambda x: x + pd.DateOffset(days=365) if pd.notna(x) and x < today else x)
    df['NDD'] = pd.to_datetime(df['NDD']).dt.date

    # 將NDD欄位的數據轉換為日期格式並去除時間
    df['NDD'] = pd.to_datetime(df['NDD']).dt.date

    # 新增 month 欄位，並限制取至小數點後兩位
    today = pd.to_datetime(datetime.today().date())  # 獲取今天的日期，轉換為日期型態並去除時間
    df['month'] = df['NDD'].apply(lambda x: round((pd.to_datetime(x) - today).days / 30.5, 2) if pd.notna(x) else np.nan)

    # 將「最近四個季度EPS1」「最近四個季度EPS2」「最近四個季度EPS3」「最近四個季度EPS4」非數字欄位視為0
    df['最近四個季度EPS1'] = pd.to_numeric(df['最近四個季度EPS1'], errors='coerce').fillna(0)
    df['最近四個季度EPS2'] = pd.to_numeric(df['最近四個季度EPS2'], errors='coerce').fillna(0)
    df['最近四個季度EPS3'] = pd.to_numeric(df['最近四個季度EPS3'], errors='coerce').fillna(0)
    df['最近四個季度EPS4'] = pd.to_numeric(df['最近四個季度EPS4'], errors='coerce').fillna(0)

    # 確保 'EPS' 欄位為數值型態
    df['EPS'] = pd.to_numeric(df['EPS'], errors='coerce')

    # 如果「EPS」欄位有值，則「Last 4Q EPS」=「EPS」
    # 否則，計算「最近四個季度EPS1」、「最近四個季度EPS2」、「最近四個季度EPS3」、「最近四個季度EPS4」的總和
    df['Last 4Q EPS'] = np.where(df['EPS'].notna(), df['EPS'], 
                                df['最近四個季度EPS1'] + df['最近四個季度EPS2'] + df['最近四個季度EPS3'] + df['最近四個季度EPS4'])

    # 新增「EPS+」欄位，當「Last 4Q EPS」大於等於0時，「EPS+」等於「Last 4Q EPS」，否則「EPS+」為0
    df['EPS+'] = np.where(df['Last 4Q EPS'] >= 0, df['Last 4Q EPS'], 0)

    # 新增「EPS+」欄位，當「Last 4Q EPS」大於等於0時，「EPS+」等於「Last 4Q EPS」，否則「EPS+」為0
    df['EPS+'] = np.where(df['Last 4Q EPS'] >= 0, df['Last 4Q EPS'], 0)

    # 只將「無法計算」替換為0，不影響其他值
    df['前1年度 配發率'] = df['前1年度 配發率'].replace('無法計算', 0)
    df['前2年度 配發率'] = df['前2年度 配發率'].replace('無法計算', 0)
    df['前3年度 配發率'] = df['前3年度 配發率'].replace('無法計算', 0)

    # 去除百分比符號並轉換為浮點數
    df['前1年度 配發率'] = df['前1年度 配發率'].str.rstrip('%').astype('float') / 100
    df['前2年度 配發率'] = df['前2年度 配發率'].str.rstrip('%').astype('float') / 100
    df['前3年度 配發率'] = df['前3年度 配發率'].str.rstrip('%').astype('float') / 100

    # 確保 '配息率' 欄位為數值型態
    df['配息率'] = pd.to_numeric(df['配息率'], errors='coerce')

    # 如果「配息率」欄位有值，則「M配息率」=「配息率」
    # 否則，計算「前1年度 配發率」、「前2年度 配發率」、「前3年度 配發率」的中位數
    df['M配息率'] = np.where(df['配息率'].notna(), df['配息率'], 
                            df[['前1年度 配發率', '前2年度 配發率', '前3年度 配發率']].median(axis=1))

    # 當「M配息率」大於1時，將其設為1
    df['M配息率'] = np.where(df['M配息率'] > 1, 1, df['M配息率'])

    # 確保 '下次配息金額' 欄位為數值型態
    df['下次配息金額'] = pd.to_numeric(df['下次配息金額'], errors='coerce')

    # 如果「下次配息金額」有值，則「預估配息」=「下次配息金額」
    # 否則，計算「預估配息」=「EPS+」x「M配息率」
    df['預估配息'] = np.where(df['下次配息金額'].notna(), df['下次配息金額'], 
                            df['EPS+'] * df['M配息率'])

    # 當「下一次除息金額」不是「無法取得資料」時，使用「下一次除息金額」作為「預估配息」的值
    df['預估配息'] = np.where(df['下一次除息金額'] != '無法取得資料', df['下一次除息金額'], df['預估配息'])

    # 如果「support」欄位有值，則「支撐」=「預估配息」/「support」
    # 如果「support」欄位沒有值，則「支撐」=「預估配息」/ 0.05
    df['支撐'] = np.where(df['support'].notna(), df['預估配息'] / df['support'], df['預估配息'] / 0.05)

    # 確保'最新收盤價'欄位是數值型態
    df['最新收盤價'] = pd.to_numeric(df['最新收盤價'], errors='coerce')

    # 新增「預期報酬」欄位
    df['預期報酬'] = df['支撐'] / df['最新收盤價'] - 1

    # 新增「預期月報酬」欄位，當month欄位不為0時計算，為0時設為NaN以避免除以0的錯誤
    df['預期月報酬'] = df.apply(lambda x: x['預期報酬'] / x['month'] if x['month'] != 0 else np.nan, axis=1)

    # 先按照「預期月報酬」欄位由大至小排序
    df = df.sort_values(by='預期月報酬', ascending=False)

    # 將資料輸出至新的 Excel 檔案
    output_file_path = r'G:\我的雲端硬碟\Horizon\python_stock\qualified_stocks_financial_data_with_estimated_payout_and_NDD.xlsx'
    with pd.ExcelWriter(output_file_path, engine='openpyxl', date_format='yyyy-mm-dd') as writer:
        df.to_excel(writer, index=False)

    # 使用 openpyxl 加載 Excel 文件，並進行文字標記
    wb = load_workbook(output_file_path)
    ws = wb.active

    # 讀取 "EPS持股" sheet 中的 "position" 欄位作為比對依據
    eps_position_df = pd.read_excel(manual_list_path, sheet_name='EPS持股')
    positions = eps_position_df['position'].astype(str).tolist()

    # 比對「股票代碼」和 "position" 欄位，匹配時將該行文字設為紅色並加粗
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        stock_code = str(row[0].value)  # 假設「股票代碼」在第一欄
        if stock_code in positions:
            for cell in row:
                cell.font = Font(color="FF0000", bold=True)  # 設置文字顏色為紅色並加粗

    # 凍結窗格
    ws.freeze_panes = 'B2'

    # 保存 Excel 文件
    wb.save(output_file_path)

    print(f"已成功將資料輸出至 {output_file_path} 並標記重複的股票代碼行文字為紅色")

# 主函數執行順序
def main():
    df = fetch_closing_prices()
    calculate_and_combine(df)

if __name__ == "__main__":
    main()
