import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import os
import openpyxl
import re  # 加入正則表達式
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.cell import MergedCell

# 營收資料抓取函數
def fetch_revenue(year, month, stock_code, mode='a'):
    republic_year = year - 1911  # 西元轉民國年
    if mode == 'a':  # 上市公司
        url = f'https://mops.twse.com.tw/nas/t21/sii/t21sc03_{republic_year}_{month}_0.html'
    elif mode == 'b':  # 上櫫公司
        url = f'https://mops.twse.com.tw/nas/t21/otc/t21sc03_{republic_year}_{month}_0.html'
    
    response = requests.get(url)
    response.encoding = 'big5'
    soup = BeautifulSoup(response.text, 'html.parser')
    tables = soup.find_all('table')
    for table in tables:
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all('td')
            if len(cells) > 0 and cells[0].get_text(strip=True) == stock_code:
                revenue = cells[2].get_text(strip=True).replace(',', '')
                return int(revenue) if revenue.isdigit() else None
    return None  # 返回 None 表示該月份無數據

# 抓取財務報表資料的共通函數 (StatementOfComprehensiveIncome 或 BalanceSheet)
def fetch_financial_data(stock_code, years, target_codes, section_id, mode, specific_quarter=None, need_url=True):
    dfs = []
    urls = [] if need_url else None  # 用來儲存每個季度的 URL，如果不需要 URL，則設為 None
    last_quarter = None  # 用來存儲最後抓取的季度

    for year in years:
        for quarter in range(1, 5):
            year_quarter = f"{year}Q{quarter}"

            if specific_quarter and year_quarter != specific_quarter:
                continue  # 如果指定了具體季度，跳過其他季度

            if mode == 'A':  # 上市公司
                url = f"https://mops.twse.com.tw/server-java/t164sb01?step=3&SYEAR={year}&file_name=tifrs-fr1-m1-ci-cr-{stock_code}-{year}Q{quarter}.html"
            elif mode == 'B':  # 上櫃公司
                url = f"https://mops.twse.com.tw/server-java/t164sb01?step=1&CO_ID={stock_code}&SYEAR={year}&SSEASON={quarter}&REPORT_ID=A"

            try:
                response = requests.get(url, timeout=5)
                if response.status_code == 200:
                    response.encoding = 'big5'
                    soup = BeautifulSoup(response.text, 'html.parser')
                    div = soup.find('div', id=section_id)  # 動態查找 section

                    if div is None:
                        continue  # 如果没有找到 div，跳过当前季度

                    table = div.find_next('table')
                    rows = table.find_all('tr')
                    data = []
                    for row in rows:
                        cols = row.find_all('td')
                        if not cols:
                            continue
                        code = cols[0].text.strip()
                        account_item = cols[1].text.strip()
                        value = cols[2].text.strip()

                        if "(" in value and ")" in value:
                            value = "-" + value.replace("(", "").replace(")", "")
                        try:
                            value = float(value.replace(',', ''))
                            if code in ['9750', '9850']:
                                value = round(value, 2)  # 9750 和 9850 保留兩位小數
                            else:
                                value = int(value)  # 其他代號轉換為整數
                        except ValueError:
                            value = None

                        if code in target_codes:
                            data.append([code, account_item, value])

                    if len(data) > 0:
                        quarter_df = pd.DataFrame(data, columns=['代號', '會計項目', year_quarter])
                        dfs.append(quarter_df)
                        if need_url:
                            urls.append(url)  # 將當前 URL 加入列表
                        last_quarter = year_quarter  # 更新最後抓取的季度

            except requests.RequestException as e:
                print(f"Failed to fetch data for {year_quarter}: {str(e)}")

    if dfs:
        final_df = dfs[0]
        for df in dfs[1:]:
            final_df = pd.merge(final_df, df, on=['代號', '會計項目'], how='outer')

        # 如果需要URL，則在資料的最後一列添加 URL 列
        if need_url:
            url_row = ['URL', ''] + urls  # 第一列為 'URL'
            final_df.loc[len(final_df)] = url_row  # 在最後一列插入URL

        return final_df, last_quarter
    else:
        return pd.DataFrame(), None  # 空的DataFrame 和 None

# 自動生成連續年份的功能
def generate_year_range(start_year, end_year):
    return list(range(start_year, end_year + 1))

# 匯入其他Excel檔案的指定工作表
def import_sheets_from_excel(file_path, sheets, writer):
    wb = load_workbook(file_path)
    for sheet_name in sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 將工作表內容轉換為 DataFrame
            data = ws.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)
            # 將 DataFrame 寫入目標文件的工作表
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def apply_formatting(src_ws, target_ws):
    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):  # 檢查是否為合併單元格
                continue  # 如果是合併單元格，跳過此單元格
            target_cell = target_ws.cell(row=cell.row, column=cell.column)
            # 單獨設置字體、填充、邊框、對齊和數字格式
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.alignment = copy(cell.alignment)
                target_cell.number_format = cell.number_format

    # 設置列寬
    for col_letter, col_dim in src_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width

def freeze_panes_in_worksheet(target_wb, sheet_name, freeze_cell):
    """
    凍結指定工作表的窗格。
    :param target_wb: openpyxl 加載的工作簿
    :param sheet_name: 工作表名稱
    :param freeze_cell: 凍結窗格的起始單元格，例如 "C3"
    """
    if sheet_name in target_wb.sheetnames:
        ws = target_wb[sheet_name]
        ws.freeze_panes = freeze_cell

# 整合多個報表資料並將結果寫入不同工作表
def combine_revenue_and_financial_data(stock_code, start_year, end_year, target_codes, mode_revenue='a', mode_financial='A'):
    # 生成連續年份範圍
    years = generate_year_range(start_year, end_year)

    # 1. 抓取營收資料
    revenue_data = {'月份': [], '當月營收': []}
    last_month = None  # 用來存儲最後抓取的月份

    for year in years:
        for month in range(1, 13):
            revenue = fetch_revenue(year, month, stock_code, mode_revenue)
            if revenue is not None:
                year_month = int(f"{year}{str(month).zfill(2)}")
                revenue_data['月份'].append(year_month)
                revenue_data['當月營收'].append(revenue)
                last_month = year_month  # 更新最後抓取的月份
            else:
                revenue_data['月份'].append("")
                revenue_data['當月營收'].append("")

    revenue_df = pd.DataFrame(revenue_data)

    # 2. 抓取財務報表資料 - 綜合損益表
    income_df, last_quarter = fetch_financial_data(stock_code, years, target_codes, 'StatementOfComprehensiveIncome', mode_financial)

    # 3. 抓取財務報表資料 - 只抓取綜合損益表成功的最後一個季度的資產負債表 (代號 3110)，且不需要URL
    if last_quarter:
        balance_sheet_df, _ = fetch_financial_data(stock_code, years, ['3110'], 'BalanceSheet', mode_financial, specific_quarter=last_quarter, need_url=False)
    else:
        balance_sheet_df = pd.DataFrame()  # 如果沒有成功抓取綜合損益表，則返回空的DataFrame

    # 4. 動態生成檔案名稱並指定保存路徑
    folder_path = r'G:\我的雲端硬碟\Horizon\python_stock\stocks'  # 指定目標資料夾
    if last_quarter and last_month:
        filename = f"{stock_code}_is_{last_quarter}_{last_month}.xlsx"
    else:
        filename = f"{stock_code}_is_combined_data.xlsx"  # 如果無資料，則使用預設名稱
    file_path = os.path.join(folder_path, filename)  # 完整檔案路徑

    # 5. 將多個報表寫入同一Excel文件，不同工作表
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # 先寫入 Financial Statements
        income_df.to_excel(writer, index=False, sheet_name='Financial Statements')

        # 寫入 Balance Sheet
        if not balance_sheet_df.empty:
            balance_sheet_df.to_excel(writer, index=False, sheet_name='Balance Sheet')

        # 再寫入 Revenue
        revenue_df.to_excel(writer, index=False, sheet_name='Revenue')

        # 匯入其他Excel檔案的指定工作表
        import_file_path = r'G:\我的雲端硬碟\Horizon\python_stock\1537_is_2024Q3_202501_formation.xlsx'
        import_sheets = ['★IS(IFRS項目)', 'breakdown', 'Financial Statements_adj']
        import_sheets_from_excel(import_file_path, import_sheets, writer)

    # 6. 複製格式到目標檔案
    source_wb = load_workbook(r'G:\我的雲端硬碟\Horizon\python_stock\1537_is_2024Q3_202501_formation.xlsx')
    target_wb = load_workbook(file_path)

    # 遍歷來源檔案中的所有工作表
    for sheet_name in source_wb.sheetnames:
        if sheet_name in target_wb.sheetnames:
            # 獲取來源工作表和目標工作表
            source_ws = source_wb[sheet_name]
            target_ws = target_wb[sheet_name]
            # 將來源表格的格式應用到目標工作表
            apply_formatting(source_ws, target_ws)

            # 對「★IS(IFRS項目)」套用凍結窗格
            if sheet_name == "★IS(IFRS項目)":
                freeze_panes_in_worksheet(target_wb, sheet_name, "C3")

    # 保存更新後的目標檔案
    target_wb.save(file_path)

    print(f"資料已成功合併並輸出到 '{file_path}'，且格式已套用至「★IS(IFRS項目)」sheet")

# 使用示例
stock_code = '1537'
start_year = 2021
end_year = 2025
target_codes = ['4000', '5000', '6000', '6500', '6900', '7100', '7010', '7020', '7050', '7060', '7000', '7900', '7950', '8000', '8200', '8300', '8500', '8610', '8710', '8720', '9750', '9850']
combine_revenue_and_financial_data(stock_code, start_year, end_year, target_codes, mode_revenue='a', mode_financial='A')
